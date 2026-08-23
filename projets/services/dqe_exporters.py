"""
Génération des exports DQE (PDF via reportlab, Excel via openpyxl) à
partir de la structure commune produite par calculer_projet_dqe.

Structure attendue pour `dqe_data` (voir dqe_calculator.calculer_projet_dqe) :
    {
        "projet": {"id", "nom", "description", "usage_batiment",
                   "nb_niveaux", "numero_devis", "date_edition"},
        "lignes": [...],
        "lots": [{"lot": <code>, "lignes": [...], "sous_total": int}, ...],
        "sous_totaux": {...},   # par catégorie (béton/coffrage/acier)
        "total_general": int,
        "montant_lettres": str,
        "devise": "FCFA",
    }

Les deux exports regroupent les lignes par LOT (façon CIMBAT) : un
récapitulatif général (un total par lot) suivi du détail de chaque lot.

`entreprise` est un dict optionnel (voir views._entreprise_export_dict)
permettant de personnaliser l'en-tête avec le logo et les coordonnées de
l'utilisateur. Absent ou vide, l'en-tête société est simplement omis.
"""

from io import BytesIO
from decimal import Decimal
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
    PageBreak, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# Nom d'affichage générique par numéro de lot (les deux sous-lots
# GROS_OEUVRE_INFRA / GROS_OEUVRE_SUPER du modèle sont fusionnés sous le
# même numéro "02" pour l'affichage, comme dans le DQE de référence).
LOT_NOMS = {
    "00": "GÉNÉRALITÉS",
    "01": "TERRASSEMENT",
    "02": "GROS ŒUVRE",
    "03": "ÉTANCHÉITÉ",
    "04": "PLOMBERIE",
    "05": "ASSAINISSEMENT",
    "06": "ÉLECTRICITÉ",
    "07": "CHARPENTE",
    "08": "COUVERTURE",
}

COULEUR_ENTETE = colors.HexColor("#2C3E50")
COULEUR_TOTAL = colors.HexColor("#ECF0F1")
COULEUR_TOTAL_GENERAL = colors.HexColor("#FFF200")


def formater_nombre(valeur) -> str:
    """Formate un nombre avec un espace comme séparateur des milliers."""
    if valeur is None:
        return "0"
    if isinstance(valeur, (int, Decimal)):
        if isinstance(valeur, Decimal) and valeur != valeur.to_integral_value():
            valeur = float(valeur)
        else:
            return f"{int(valeur):,}".replace(",", " ")
    return f"{valeur:,.4f}".replace(",", " ").rstrip("0").rstrip(".")


def formater_date(date_iso: str) -> str:
    """'2026-08-19' -> '19/08/2026'. Retourne la valeur brute si non parsable."""
    try:
        return datetime.strptime(date_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return date_iso or ""


def _numero_lot(code_lot: str) -> str:
    """'lot_02_gros_oeuvre_superstructure' -> '02'."""
    parts = code_lot.split("_")
    if len(parts) >= 2 and parts[0] == "lot":
        return parts[1]
    return "99"


def regrouper_par_lot(dqe_data: dict) -> list:
    """
    Fusionne dqe_data["lots"] par numéro de lot (deux sous-lots Gros
    Œuvre Infra/Super comptent comme un seul "LOT 02") et retourne une
    liste ordonnée de dicts :
        [{"numero": "00", "nom": "GÉNÉRALITÉS",
          "label": "LOT 00 — GÉNÉRALITÉS",
          "lignes": [...], "sous_total": int}, ...]
    """
    groupes = {}
    for lot_entry in dqe_data.get("lots", []):
        numero = _numero_lot(lot_entry["lot"])
        groupe = groupes.setdefault(numero, {"lignes": [], "sous_total": 0})
        groupe["lignes"].extend(lot_entry["lignes"])
        groupe["sous_total"] += lot_entry["sous_total"]

    resultat = []
    for numero in sorted(groupes.keys()):
        nom = LOT_NOMS.get(numero, numero)
        resultat.append({
            "numero": numero,
            "nom": nom,
            "label": f"LOT {numero} — {nom}",
            "lignes": groupes[numero]["lignes"],
            "sous_total": groupes[numero]["sous_total"],
        })
    return resultat


def _entreprise_champ(entreprise: dict, cle: str) -> str:
    if not entreprise:
        return ""
    return (entreprise.get(cle) or "").strip()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def exporter_dqe_pdf(dqe_data: dict, entreprise: dict = None) -> BytesIO:
    """
    Génère le document PDF du DQE : en-tête société (logo + coordonnées,
    si fournis), récapitulatif général par lot, puis détail par lot.
    """
    lots = regrouper_par_lot(dqe_data)
    projet = dqe_data.get("projet", {})

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DQETitle', parent=styles['Heading1'], fontName='Helvetica-Bold',
        fontSize=16, spaceAfter=4, alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'DQESubtitle', parent=styles['Normal'], fontName='Helvetica',
        fontSize=10, alignment=1, textColor=COULEUR_ENTETE,
    )
    lots_line_style = ParagraphStyle(
        'DQELotsLine', parent=styles['Normal'], fontName='Helvetica-Oblique',
        fontSize=8, alignment=1, textColor=colors.HexColor("#555555"),
    )
    normal_style = ParagraphStyle(
        'DQENormal', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14,
    )
    small_style = ParagraphStyle(
        'DQESmall', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11,
    )
    bold_style = ParagraphStyle(
        'DQEBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=14,
    )
    header_style = ParagraphStyle(
        'DQEHeader', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=9, textColor=colors.whitesmoke,
    )
    section_style = ParagraphStyle(
        'DQESection', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=12, textColor=colors.white, backColor=COULEUR_ENTETE,
        leftIndent=6, spaceBefore=0, spaceAfter=0, leading=20,
    )

    # -- En-tête entreprise (logo + coordonnées), si renseignée --------
    logo_path = _entreprise_champ(entreprise, "logo_path")
    nom_societe = _entreprise_champ(entreprise, "nom")
    if logo_path or nom_societe:
        lignes_societe = []
        if nom_societe:
            lignes_societe.append(Paragraph(f"<b>{nom_societe}</b>", small_style))
        siege = _entreprise_champ(entreprise, "siege_social")
        if siege:
            lignes_societe.append(Paragraph(f"Siège social : {siege}", small_style))
        contact_bits = [
            b for b in [
                _entreprise_champ(entreprise, "telephone"),
                _entreprise_champ(entreprise, "email"),
            ] if b
        ]
        if contact_bits:
            lignes_societe.append(Paragraph(" - ".join(contact_bits), small_style))
        rccm = _entreprise_champ(entreprise, "rccm")
        cc = _entreprise_champ(entreprise, "cc")
        cb = _entreprise_champ(entreprise, "cb")
        legal_bits = [
            b for b in [
                f"RCCM {rccm}" if rccm else "",
                f"CC N° {cc}" if cc else "",
                f"CB N° {cb}" if cb else "",
            ] if b
        ]
        if legal_bits:
            lignes_societe.append(Paragraph(" - ".join(legal_bits), small_style))

        logo_cell = ""
        if logo_path:
            try:
                logo_cell = Image(logo_path, width=45 * mm, height=25 * mm, kind="proportional")
            except Exception:
                logo_cell = ""

        header_table = Table(
            [[logo_cell, lignes_societe]],
            colWidths=[100, 415],
        )
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 10))

    # -- Titre / devis / projet -----------------------------------------
    story.append(Paragraph("DEVIS QUANTITATIF ET ESTIMATIF (DQE)", title_style))
    entete_bits = [f"Devis N° {projet.get('numero_devis', '')}"]
    if projet.get("date_edition"):
        entete_bits.append(f"Édition du {formater_date(projet['date_edition'])}")
    story.append(Paragraph(" — ".join(entete_bits), subtitle_style))
    titre_projet = projet.get("nom", "")
    if projet.get("description"):
        titre_projet += f" — {projet['description']}"
    story.append(Paragraph(f"<b>{titre_projet}</b>", subtitle_style))
    if lots:
        story.append(Paragraph(" · ".join(l["label"] for l in lots), lots_line_style))
    story.append(Spacer(1, 14))

    # -- Récapitulatif général -------------------------------------------
    story.append(Paragraph("RÉCAPITULATIF GÉNÉRAL", section_style))
    story.append(Spacer(1, 6))

    recap_data = [[
        Paragraph("N°", header_style),
        Paragraph("DÉSIGNATION", header_style),
        Paragraph(f"MONTANT ({dqe_data.get('devise', 'FCFA')})", header_style),
    ]]
    for lot in lots:
        recap_data.append([
            Paragraph(lot["numero"], normal_style),
            Paragraph(f"TOTAL {lot['label']} (HT)", normal_style),
            Paragraph(formater_nombre(lot["sous_total"]), normal_style),
        ])
    recap_data.append([
        "", Paragraph("<b>TOTAL HT</b>", bold_style),
        Paragraph(f"<b>{formater_nombre(dqe_data['total_general'])}</b>", bold_style),
    ])
    recap_data.append([
        "", Paragraph("<b>TOTAL GÉNÉRAL HT</b>", bold_style),
        Paragraph(f"<b>{formater_nombre(dqe_data['total_general'])}</b>", bold_style),
    ])

    recap_table = Table(recap_data, colWidths=[35, 355, 125])
    recap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COULEUR_ENTETE),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -3), 0.5, colors.lightgrey),
        ('LINEABOVE', (0, -2), (-1, -2), 1, COULEUR_ENTETE),
        ('BACKGROUND', (0, -2), (-1, -2), COULEUR_TOTAL),
        ('BACKGROUND', (0, -1), (-1, -1), COULEUR_TOTAL_GENERAL),
    ]))
    story.append(recap_table)

    montant_lettres = dqe_data.get("montant_lettres")
    if montant_lettres:
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Arrêté le présent devis à la somme de ({dqe_data.get('devise', 'FCFA')}) : "
            f"<i>{montant_lettres}</i>",
            small_style,
        ))

    # -- Détail par lot ----------------------------------------------------
    story.append(PageBreak())
    detail_headers = ["Désignation", "Unité", "Quantité", "P.U. (FCFA)", "Montant (FCFA)"]
    col_widths_detail = [225, 45, 65, 90, 90]

    for lot in lots:
        if not lot["lignes"]:
            continue

        data = [[Paragraph(h, header_style) for h in detail_headers]]
        for ligne in lot["lignes"]:
            data.append([
                Paragraph(ligne["designation"], normal_style),
                Paragraph(ligne["unite"], normal_style),
                Paragraph(formater_nombre(ligne["quantite"]), normal_style),
                Paragraph(formater_nombre(ligne["prix_unitaire"]), normal_style),
                Paragraph(formater_nombre(ligne["montant"]), normal_style),
            ])
        data.append([
            Paragraph(f"<b>TOTAL {lot['label']} (HT)</b>", bold_style), "", "", "",
            Paragraph(f"<b>{formater_nombre(lot['sous_total'])}</b>", bold_style),
        ])

        t = Table(data, colWidths=col_widths_detail)
        n_rows = len(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COULEUR_ENTETE),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, n_rows - 2), 0.5, colors.lightgrey),
            ('SPAN', (0, n_rows - 1), (3, n_rows - 1)),
            ('LINEABOVE', (0, n_rows - 1), (-1, n_rows - 1), 1, COULEUR_ENTETE),
            ('BACKGROUND', (0, n_rows - 1), (-1, n_rows - 1), COULEUR_TOTAL),
        ]))

        # Le titre du lot reste collé à son tableau (pas de titre orphelin
        # en bas de page) ; le tableau lui-même peut se scinder normalement
        # sur plusieurs pages s'il est long.
        story.append(KeepTogether([Paragraph(lot["label"], section_style), Spacer(1, 6), t]))
        story.append(Spacer(1, 14))

    # -- Signature ----------------------------------------------------------
    story.append(Spacer(1, 30))
    story.append(Paragraph("<b>Pour l'ingénieur structure responsable :</b>", normal_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Date de validation : ________________________", normal_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Signature & Cachet :", normal_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _feuille_nom_unique(wb: Workbook, base: str) -> str:
    """openpyxl limite les noms de feuille à 31 caractères et refuse les doublons."""
    base = base[:31] or "Lot"
    nom = base
    i = 2
    while nom in wb.sheetnames:
        suffixe = f" ({i})"
        nom = base[: 31 - len(suffixe)] + suffixe
        i += 1
    return nom


def exporter_dqe_excel(dqe_data: dict, entreprise: dict = None) -> BytesIO:
    """
    Génère le classeur Excel du DQE : une feuille "Récapitulatif" (en-tête
    société + total par lot) puis une feuille de détail par lot.
    """
    lots = regrouper_par_lot(dqe_data)
    projet = dqe_data.get("projet", {})

    font_title = Font(name="Calibri", size=16, bold=True, color="2C3E50")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_normal = Font(name="Calibri", size=11)
    font_small = Font(name="Calibri", size=9, color="555555")

    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_total = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    fill_total_general = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'),
    )

    wb = Workbook()

    # ---- Feuille Récapitulatif ------------------------------------------
    ws = wb.active
    ws.title = "Récapitulatif"
    ws.views.sheetView[0].showGridLines = True

    row = 1
    logo_path = _entreprise_champ(entreprise, "logo_path")
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.height = 70
            img.width = 120
            ws.add_image(img, "A1")
            row = 6
        except Exception:
            pass

    nom_societe = _entreprise_champ(entreprise, "nom")
    if nom_societe:
        ws.cell(row=row, column=1, value=nom_societe).font = font_bold
        row += 1
        for cle, prefixe in [("siege_social", "Siège social : "), ("telephone", "Tél : "), ("email", "Email : ")]:
            valeur = _entreprise_champ(entreprise, cle)
            if valeur:
                ws.cell(row=row, column=1, value=f"{prefixe}{valeur}").font = font_small
                row += 1
        row += 1

    ws.cell(row=row, column=1, value="DEVIS QUANTITATIF ET ESTIMATIF (DQE)").font = font_title
    row += 1
    ws.cell(row=row, column=1, value=f"Devis N° {projet.get('numero_devis', '')} — Édition du {formater_date(projet.get('date_edition', ''))}").font = font_normal
    row += 1
    titre_projet = projet.get("nom", "")
    if projet.get("description"):
        titre_projet += f" — {projet['description']}"
    ws.cell(row=row, column=1, value=titre_projet).font = font_bold
    row += 2

    headers = ["N°", "Désignation", f"Montant ({dqe_data.get('devise', 'FCFA')})"]
    header_row = row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    row += 1

    for lot in lots:
        ws.cell(row=row, column=1, value=lot["numero"]).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=f"TOTAL {lot['label']} (HT)").border = thin_border
        m_cell = ws.cell(row=row, column=3, value=lot["sous_total"])
        m_cell.number_format = "#,##0"
        m_cell.alignment = Alignment(horizontal="right")
        m_cell.border = thin_border
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    tot_ht_lbl = ws.cell(row=row, column=1, value="TOTAL HT")
    tot_ht_lbl.font = font_bold
    tot_ht_lbl.alignment = Alignment(horizontal="right")
    tot_ht_lbl.fill = fill_total
    tot_ht_val = ws.cell(row=row, column=3, value=dqe_data["total_general"])
    tot_ht_val.font = font_bold
    tot_ht_val.number_format = "#,##0"
    tot_ht_val.alignment = Alignment(horizontal="right")
    tot_ht_val.fill = fill_total
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    tot_lbl = ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL HT")
    tot_lbl.font = font_bold
    tot_lbl.alignment = Alignment(horizontal="right")
    tot_lbl.fill = fill_total_general
    tot_val = ws.cell(row=row, column=3, value=dqe_data["total_general"])
    tot_val.font = font_bold
    tot_val.number_format = "#,##0"
    tot_val.alignment = Alignment(horizontal="right")
    tot_val.fill = fill_total_general
    row += 2

    if dqe_data.get("montant_lettres"):
        ws.cell(
            row=row, column=1,
            value=f"Arrêté le présent devis à la somme de : {dqe_data['montant_lettres']}",
        ).font = font_small

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20

    # ---- Une feuille de détail par lot ------------------------------------
    for lot in lots:
        if not lot["lignes"]:
            continue
        ws_lot = wb.create_sheet(_feuille_nom_unique(wb, lot["label"]))
        ws_lot.views.sheetView[0].showGridLines = True

        ws_lot.cell(row=1, column=1, value=lot["label"]).font = font_title
        ws_lot.row_dimensions[1].height = 22

        headers = ["Désignation", "Unité", "Quantité", "Prix Unitaire (FCFA)", "Montant (FCFA)"]
        start_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws_lot.cell(row=start_row, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_lot.row_dimensions[start_row].height = 24

        current_row = start_row + 1
        for ligne in lot["lignes"]:
            ws_lot.cell(row=current_row, column=1, value=ligne["designation"]).font = font_normal
            ws_lot.cell(row=current_row, column=1).border = thin_border

            ws_lot.cell(row=current_row, column=2, value=ligne["unite"]).font = font_normal
            ws_lot.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
            ws_lot.cell(row=current_row, column=2).border = thin_border

            q_cell = ws_lot.cell(row=current_row, column=3, value=float(ligne["quantite"]))
            q_cell.font = font_normal
            q_cell.number_format = "0.####"
            q_cell.alignment = Alignment(horizontal="right")
            q_cell.border = thin_border

            pu_cell = ws_lot.cell(row=current_row, column=4, value=int(ligne["prix_unitaire"]))
            pu_cell.font = font_normal
            pu_cell.number_format = "#,##0"
            pu_cell.alignment = Alignment(horizontal="right")
            pu_cell.border = thin_border

            m_cell = ws_lot.cell(row=current_row, column=5, value=int(ligne["montant"]))
            m_cell.font = font_normal
            m_cell.number_format = "#,##0"
            m_cell.alignment = Alignment(horizontal="right")
            m_cell.border = thin_border

            current_row += 1

        ws_lot.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        lbl_cell = ws_lot.cell(row=current_row, column=1, value=f"TOTAL {lot['label']} (HT)")
        lbl_cell.font = font_bold
        lbl_cell.alignment = Alignment(horizontal="right")
        lbl_cell.fill = fill_total

        val_cell = ws_lot.cell(row=current_row, column=5, value=lot["sous_total"])
        val_cell.font = font_bold
        val_cell.number_format = "#,##0"
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.fill = fill_total

        ws_lot.column_dimensions["A"].width = 45
        ws_lot.column_dimensions["B"].width = 10
        ws_lot.column_dimensions["C"].width = 15
        ws_lot.column_dimensions["D"].width = 20
        ws_lot.column_dimensions["E"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer