from decimal import Decimal
from io import BytesIO
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase
from openpyxl import load_workbook

from projets.models import Projet, ElementStructurel, PosteComplementaire
from projets.services.dqe_calculator import calculer_element_dqe, calculer_projet_dqe
from projets.services.dqe_exporters import exporter_dqe_pdf, exporter_dqe_excel

class DQECalculatorTestCase(TestCase):
    def setUp(self):
        self.projet = Projet.objects.create(
            nom="Immeuble Test R+1",
            usage_batiment="habitation",
            nb_niveaux=2
        )
        # 1. Poteau (avec cote_cm inclus pour la compatibilité DQE)
        self.poteau = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POTEAU,
            identifiant="P1",
            hauteur_poteau=3.0,
            resultat_calcul={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            resultat_valide={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            statut=ElementStructurel.Statut.VALIDE
        )
        # 2. Poutre
        self.poutre = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POUTRE,
            identifiant="PT1",
            portee=5.0,
            resultat_calcul={"largeur_cm": 20, "hauteur_cm": 40},
            resultat_valide={"largeur_cm": 20, "hauteur_cm": 40},
            statut=ElementStructurel.Statut.VALIDE
        )
        # 3. Semelle
        self.semelle = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.SEMELLE,
            identifiant="S1",
            resultat_calcul={"cote_cm": 150, "hauteur_cm": 40},
            resultat_valide={"cote_cm": 150, "hauteur_cm": 40},
            statut=ElementStructurel.Statut.VALIDE
        )

    def test_calculer_element_poteau_ratio(self):
        lignes = calculer_element_dqe(self.poteau, {})
        # Doit générer 3 lignes (Béton, Coffrage, Acier)
        self.assertEqual(len(lignes), 3)
        
        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        # volume_beton = 0.20 * 0.20 * 3.0 = 0.12 m3
        self.assertEqual(beton["quantite"], 0.12)
        # surf_coffrage = 4 * 0.20 * 3.0 = 2.4 m2
        self.assertEqual(coffrage["quantite"], 2.4)
        # poids_acier = 0.12 * 125 = 15.0 kg
        self.assertEqual(acier["quantite"], 15.0)

    def test_calculer_element_poteau_poids_moteur(self):
        self.poteau.resultat_valide = {
            "cote_cm": 20,
            "largeur_cm": 20,
            "profondeur_cm": 20,
            "poids_acier_total_kg": 15.5,
            "poids_acier_kg": 15.5
        }
        self.poteau.save()

        lignes = calculer_element_dqe(self.poteau, {})
        acier = next(l for l in lignes if l["categorie"] == "ACIER")
        self.assertEqual(acier["quantite"], 15.5)

    def test_calculer_element_poutre(self):
        lignes = calculer_element_dqe(self.poutre, {})
        self.assertEqual(len(lignes), 3)

        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        self.assertEqual(beton["quantite"], 0.40)
        self.assertEqual(coffrage["quantite"], 5.0)
        # poids_acier = 0.40 * 150 = 60 kg
        self.assertEqual(acier["quantite"], 60.0)

    def test_calculer_element_semelle(self):
        lignes = calculer_element_dqe(self.semelle, {})
        self.assertEqual(len(lignes), 3)

        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        self.assertEqual(beton["quantite"], 0.90)
        self.assertEqual(coffrage["quantite"], 2.4)
        # poids_acier = 0.90 * 50 = 45 kg
        self.assertEqual(acier["quantite"], 45.0)

    def test_calculer_projet_exclut_non_valides(self):
        self.poutre.statut = ElementStructurel.Statut.PROPOSE
        self.poutre.save()
        
        self.semelle.statut = ElementStructurel.Statut.MODIFIE
        self.semelle.save()

        dqe_data = calculer_projet_dqe(self.projet)
        
        reperes = [l["repere"] for l in dqe_data["lignes"]]
        self.assertIn("P1", reperes)
        self.assertNotIn("PT1", reperes)
        self.assertNotIn("S1", reperes)

    def test_calculer_projet_dqe_global_avec_main_doeuvre(self):
        PosteComplementaire.objects.create(
            projet=self.projet,
            lot=PosteComplementaire.Lot.ELECTRICITE,
            designation="Terrassement fouilles",
            unite="m³",
            quantite=15.0,
            prix_unitaire=5000.0
        )

        dqe_data = calculer_projet_dqe(self.projet)

        mo_line = next(l for l in dqe_data["lignes"] if l["type_element"] == "MAIN_DOEUVRE")
        self.assertEqual(mo_line["designation"], "Terrassement fouilles")
        self.assertEqual(mo_line["montant"], 75000)

        # Un sous-total plat "main_doeuvre" mélangerait main d'œuvre et
        # gros œuvre dans un seul total sans distinction de lot ; les
        # postes de main d'œuvre sont donc comptabilisés dans le
        # sous-total de LEUR lot (voir calculer_projet_dqe).
        lot_electricite = next(
            l for l in dqe_data["lots"] if l["lot"] == PosteComplementaire.Lot.ELECTRICITE.value
        )
        self.assertEqual(lot_electricite["sous_total"], 75000)

    def test_calculer_element_dalle(self):
        dalle = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.DALLE,
            identifiant="D1",
            portee=4.0,
            surface_m2=50.0,
            resultat_calcul={"epaisseur_cm": 15},
            resultat_valide={"epaisseur_cm": 15},
            statut=ElementStructurel.Statut.VALIDE
        )
        lignes = calculer_element_dqe(dalle, {})
        self.assertEqual(len(lignes), 3)

        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        self.assertEqual(beton["quantite"], 7.5)
        self.assertEqual(coffrage["quantite"], 50.0)
        self.assertEqual(acier["quantite"], 637.5)

    def test_calculer_element_semelle_filante_poids_moteur(self):
        sf = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.SEMELLE_FILANTE,
            identifiant="SF1",
            longueur_m=10.0,
            resultat_calcul={
                "largeur_cm": 50,
                "hauteur_cm": 30,
                "acier_transversal_cm2_ml": 4.0,
                "acier_repartition_cm2_ml": 2.0
            },
            resultat_valide={
                "largeur_cm": 50,
                "hauteur_cm": 30,
                "acier_transversal_cm2_ml": 4.0,
                "acier_repartition_cm2_ml": 2.0
            },
            statut=ElementStructurel.Statut.VALIDE
        )
        lignes = calculer_element_dqe(sf, {})
        self.assertEqual(len(lignes), 3)

        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        self.assertEqual(beton["quantite"], 1.5)
        self.assertEqual(coffrage["quantite"], 6.0)
        self.assertEqual(acier["quantite"], 47.1)  # 6.0 cm2/ml * 10m * 7.85 kg/m/cm2 = 47.1 kg

    def test_calculer_element_semelle_filante_ratio(self):
        sf = ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.SEMELLE_FILANTE,
            identifiant="SF1",
            longueur_m=10.0,
            resultat_calcul={
                "largeur_cm": 50,
                "hauteur_cm": 30
            },
            resultat_valide={
                "largeur_cm": 50,
                "hauteur_cm": 30
            },
            statut=ElementStructurel.Statut.VALIDE
        )
        lignes = calculer_element_dqe(sf, {})
        self.assertEqual(len(lignes), 3)

        beton = next(l for l in lignes if l["categorie"] == "BETON")
        coffrage = next(l for l in lignes if l["categorie"] == "COFFRAGE")
        acier = next(l for l in lignes if l["categorie"] == "ACIER")

        self.assertEqual(beton["quantite"], 1.5)
        self.assertEqual(coffrage["quantite"], 6.0)
        self.assertEqual(acier["quantite"], 75.0)  # 1.5 m3 * 50 kg/m3 = 75.0 kg

    def test_calculer_projet_dqe_global_cinq_elements(self):
        # On ajoute les 2 types Phase 2 manquants au projet du setUp
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.DALLE,
            identifiant="D1",
            portee=4.0,
            surface_m2=50.0,
            resultat_calcul={"epaisseur_cm": 15},
            resultat_valide={"epaisseur_cm": 15},
            statut=ElementStructurel.Statut.VALIDE
        )
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.SEMELLE_FILANTE,
            identifiant="SF1",
            longueur_m=10.0,
            resultat_calcul={"largeur_cm": 50, "hauteur_cm": 30},
            resultat_valide={"largeur_cm": 50, "hauteur_cm": 30},
            statut=ElementStructurel.Statut.VALIDE
        )

        dqe_data = calculer_projet_dqe(self.projet)

        # Vérification des sous-totaux par catégorie
        # Béton : 0.12 (poteau) + 0.40 (poutre) + 0.90 (semelle) + 7.50 (dalle) + 1.50 (semelle filante) = 10.42 m3
        # Cost : 10.42 * 100 000 = 1 042 000 FCFA
        self.assertEqual(dqe_data["sous_totaux"]["beton"], 1042000)

        # Coffrage : 2.4 (poteau) + 5.0 (poutre) + 2.4 (semelle) + 50.0 (dalle) + 6.0 (semelle filante) = 65.8 m2
        # Cost : 65.8 * 12 000 = 789 600 FCFA
        self.assertEqual(dqe_data["sous_totaux"]["coffrage"], 789600)

        # Acier : 15.0 (poteau) + 60.0 (poutre) + 45.0 (semelle) + 637.5 (dalle) + 75.0 (semelle filante) = 832.5 kg
        # Cost : 832.5 * 800 = 666 000 FCFA
        self.assertEqual(dqe_data["sous_totaux"]["acier"], 666000)

        # Total Général : 1 042 000 + 789 600 + 666 000 = 2 497 600 FCFA
        self.assertEqual(dqe_data["total_general"], 2497600)


class DQEExportersTestCase(TestCase):
    def setUp(self):
        ligne_poteau = {
            "element_id": 1,
            "repere": "P1",
            "type_element": "POTEAU",
            "designation": "Béton armé — Poteau P1",
            "categorie": "BETON",
            "unite": "m³",
            "quantite": 0.12,
            "prix_unitaire": 100000,
            "montant": 12000
        }
        self.dqe_data = {
            "projet": {"id": 1, "nom": "Immeuble R+1"},
            "lignes": [ligne_poteau],
            "lots": [
                {
                    "lot": "lot_02_gros_oeuvre_superstructure",
                    "lignes": [ligne_poteau],
                    "sous_total": 12000,
                }
            ],
            "sous_totaux": {
                "beton": 12000,
                "coffrage": 0,
                "acier": 0,
            },
            "total_general": 12000,
            "montant_lettres": "douze mille",
            "devise": "FCFA"
        }

    def test_exporter_pdf(self):
        buffer = exporter_dqe_pdf(self.dqe_data)
        self.assertIsInstance(buffer, BytesIO)
        self.assertTrue(buffer.getvalue().startswith(b"%PDF"))

    def test_exporter_excel(self):
        buffer = exporter_dqe_excel(self.dqe_data)
        self.assertIsInstance(buffer, BytesIO)

        wb = load_workbook(buffer)
        # La feuille "Récapitulatif" (total par lot) remplace l'ancienne
        # feuille unique "DQE" ; le détail est désormais réparti sur une
        # feuille par lot (voir dqe_exporters.exporter_dqe_excel).
        self.assertIn("Récapitulatif", wb.sheetnames)
        ws = wb["Récapitulatif"]
        self.assertEqual(ws["A1"].value, "DEVIS QUANTITATIF ET ESTIMATIF (DQE)")

        self.assertIn("LOT 02 — GROS ŒUVRE", wb.sheetnames)


class DQEAPITestCase(APITestCase):
    def setUp(self):
        self.projet = Projet.objects.create(
            nom="Projet API Test",
            usage_batiment="bureau",
            nb_niveaux=3
        )
        self.url = reverse("projet-generer-dqe", kwargs={"pk": self.projet.id})

    def test_generer_dqe_sans_elements_echoue(self):
        # Aucun élément dans le projet
        response = self.client.get(f"{self.url}?export=pdf")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["erreur"], "Le projet ne contient aucun élément structurel.")

    def test_generer_dqe_avec_elements_non_valides_echoue(self):
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POTEAU,
            identifiant="P1",
            hauteur_poteau=3.0,
            statut=ElementStructurel.Statut.PROPOSE
        )
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("erreur", response.data)
        self.assertIn("P1", response.data["elements_en_attente"])

    def test_generer_dqe_format_invalide_echoue(self):
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POTEAU,
            identifiant="P1",
            hauteur_poteau=3.0,
            resultat_calcul={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            resultat_valide={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            statut=ElementStructurel.Statut.VALIDE
        )
        response = self.client.get(f"{self.url}?export=word")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("erreur", response.data)

    def test_generer_dqe_pdf_succes(self):
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POTEAU,
            identifiant="P1",
            hauteur_poteau=3.0,
            resultat_calcul={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            resultat_valide={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            statut=ElementStructurel.Statut.VALIDE
        )
        response = self.client.get(f"{self.url}?export=pdf")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_generer_dqe_excel_succes(self):
        ElementStructurel.objects.create(
            projet=self.projet,
            type_element=ElementStructurel.TypeElement.POTEAU,
            identifiant="P1",
            hauteur_poteau=3.0,
            resultat_calcul={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            resultat_valide={"cote_cm": 20, "largeur_cm": 20, "profondeur_cm": 20},
            statut=ElementStructurel.Statut.VALIDE
        )
        # Test avec GET
        response = self.client.get(f"{self.url}?export=excel")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )